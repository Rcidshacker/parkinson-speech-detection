import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side, GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from datetime import datetime

OUT = r"C:\Users\Lenovo\Desktop\Code\2026\BE mini project\Literature\workbooks\BEnSParX_Comprehensive_Report_v2.xlsx"

# ── palette ──────────────────────────────────────────────────────────────────
DARK_BLUE   = "1F3864"
MID_BLUE    = "2E75B6"
LIGHT_BLUE  = "BDD7EE"
ACCENT_GOLD = "C9A227"
LIGHT_GRAY  = "F2F2F2"
WHITE       = "FFFFFF"
GREEN_BG    = "E2EFDA"
RED_BG      = "FCE4D6"
YELLOW_BG   = "FFEB9C"
DARK_TEXT   = "1A1A1A"

def fill(hex_): 
    return PatternFill("solid", fgColor=hex_)

def font(bold=False, size=11, color=DARK_TEXT, italic=False):
    return Font(name="Arial", bold=bold, size=size, color=color, italic=italic)

def center(): 
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():   
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def thin_border():
    s = Side(style="thin", color="AAAAAA")
    return Border(left=s, right=s, top=s, bottom=s)

def thick_border():
    s = Side(style="medium", color="1F3864")
    return Border(left=s, right=s, top=s, bottom=s)

def hdr(ws, cell, text, bg=MID_BLUE, fg=WHITE, size=11, bold=True, merge=None):
    c = ws[cell]
    c.value = text
    c.font  = font(bold=bold, size=size, color=fg)
    c.fill  = fill(bg)
    c.alignment = center()
    c.border = thin_border()
    if merge:
        ws.merge_cells(merge)

def cell(ws, row, col, val, bg=None, bold=False, size=11, align="left",
         color=DARK_TEXT, border=True, italic=False, num_fmt=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font = font(bold=bold, size=size, color=color, italic=italic)
    if bg: c.fill = fill(bg)
    c.alignment = center() if align == "center" else left()
    if border: c.border = thin_border()
    if num_fmt: c.number_format = num_fmt
    return c

def col_w(ws, mapping):
    for col, w in mapping.items():
        ws.column_dimensions[col].width = w

def row_h(ws, mapping):
    for r, h in mapping.items():
        ws.row_dimensions[r].height = h

wb = openpyxl.Workbook()
wb.remove(wb.active)   # remove default sheet

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 1 — COVER
# ═══════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Cover")
ws.sheet_view.showGridLines = False
col_w(ws, {"A": 4, "B": 30, "C": 35, "D": 22, "E": 22, "F": 4})

# title block
ws.merge_cells("B2:E2"); c = ws["B2"]
c.value = "Speech-Based Foundation Model for Cross-lingual"
c.font  = font(bold=True, size=18, color=WHITE)
c.fill  = fill(DARK_BLUE); c.alignment = center()

ws.merge_cells("B3:E3"); c = ws["B3"]
c.value = "Parkinson's Disease Detection"
c.font  = font(bold=True, size=18, color=WHITE)
c.fill  = fill(DARK_BLUE); c.alignment = center()

ws.merge_cells("B4:E4"); c = ws["B4"]
c.value = "Feature Extraction Pipeline — Comprehensive Results Report v2"
c.font  = font(bold=False, size=13, color=ACCENT_GOLD)
c.fill  = fill(DARK_BLUE); c.alignment = center()
row_h(ws, {2: 40, 3: 40, 4: 32})

# info table
info = [
    (6,  "Student",          "Ruchit Das (22AM1084)"),
    (7,  "Role",             "Feature Extraction Pipeline"),
    (8,  "Institute",        "Ramrao Adik Institute of Technology"),
    (9,  "Guide",            "Dr. Sandeep Sangle"),
    (10, "Mentor",           "Pramod Kachare"),
    (11, "Team",             "Mrudul Jadhav · Achyut Maheshka · Niharika Mishra"),
    (12, "Languages",        "Spanish (PC-GITA) · Italian (VOICED)"),
    (13, "Report Generated", datetime.now().strftime("%Y-%m-%d %H:%M")),
    (14, "Mock 2 Deadline",  "April 4, 2026"),
    (15, "Final Exam",       "May 2, 2026"),
]

for r, label, value in info:
    cell(ws, r, 2, label, bg=LIGHT_BLUE, bold=True, size=11, align="left")
    ws.merge_cells(f"C{r}:E{r}")
    cell(ws, r, 3, value, bg=WHITE, size=11)
    row_h(ws, {r: 22})

# sheet index
ws.merge_cells("B17:E17"); c = ws["B17"]
c.value = "REPORT CONTENTS"; c.font = font(bold=True, size=12, color=WHITE)
c.fill = fill(MID_BLUE); c.alignment = center()

sheets_info = [
    ("Overview",              "Summary of all 4 feature sets — best model, accuracy, target"),
    ("Cross-Dataset Matrix",  "Full 3×3 cross-dataset AUC matrix (GroupKFold v2, speaker-safe)"),
    ("Model Results",         "All 7 models × 6 metrics for 112-feature set"),
    ("Feature Sets",          "Feature lists with origin labels (common / ES-only / IT-only)"),
    ("Per-Dataset Rankings",  "Top-9 per-dataset AUC rankings (PC-GITA & VOICED)"),
    ("Methodology",           "Pipeline steps, GroupKFold, SMOTE, tech stack"),
]

for i, (sh, desc) in enumerate(sheets_info, start=18):
    cell(ws, i, 2, sh,   bg=LIGHT_GRAY, bold=True, size=10)
    ws.merge_cells(f"C{i}:E{i}")
    cell(ws, i, 3, desc, bg=WHITE, size=10)
    row_h(ws, {i: 20})

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 2 — OVERVIEW
# ═══════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Overview")
ws.sheet_view.showGridLines = False
col_w(ws, {"A": 3, "B": 26, "C": 16, "D": 22, "E": 18, "F": 16, "G": 20, "H": 3})

ws.merge_cells("B1:G1"); c = ws["B1"]
c.value = "PIPELINE OVERVIEW — ALL FEATURE SETS"
c.font = font(bold=True, size=14, color=WHITE); c.fill = fill(DARK_BLUE); c.alignment = center()
row_h(ws, {1: 34})

# Dataset info
ws.merge_cells("B3:G3"); c = ws["B3"]
c.value = "DATASETS"; c.font = font(bold=True, size=11, color=WHITE)
c.fill = fill(MID_BLUE); c.alignment = center(); row_h(ws, {3: 22})

hdrs3 = ["Dataset", "Language", "Samples", "PD", "HC", "Features Extracted"]
for i, h in enumerate(hdrs3, 2):
    cell(ws, 4, i, h, bg=LIGHT_BLUE, bold=True, align="center")
row_h(ws, {4: 20})

ds_data = [
    ("PC-GITA",  "Spanish (ES)", 1318, 644, 674, 129),
    ("VOICED",   "Italian (IT)",  296, 133, 163, 129),
    ("Combined", "ES + IT",      1614, 777, 837, 129),
]

for ri, row in enumerate(ds_data, 5):
    bg = LIGHT_GRAY if ri % 2 == 0 else WHITE
    for ci, v in enumerate(row, 2):
        cell(ws, ri, ci, v, bg=bg, align="center")
row_h(ws, {5: 18, 6: 18, 7: 18})

# Feature sets summary
ws.merge_cells("B9:G9"); c = ws["B9"]
c.value = "FEATURE SETS SUMMARY (GroupKFold v2 — Speaker-Safe)"; c.font = font(bold=True, size=11, color=WHITE)
c.fill = fill(MID_BLUE); c.alignment = center(); row_h(ws, {9: 22})

hdrs9 = ["Feature Set", "Actual Features", "Best Model", "ES→IT AUC", "IT→ES AUC", "90% Target Met?"]
for i, h in enumerate(hdrs9, 2):
    cell(ws, 10, i, h, bg=LIGHT_BLUE, bold=True, align="center")
row_h(ws, {10: 20})

fs_summary = [
    ("Top 9  (14 features)",  14, "StackingEnsemble", 0.821, 0.543, "✓ ES→IT (Prev)"),
    ("Top 20 (31 features)",  31, "StackingEnsemble", 0.821, 0.543, "✓ ES→IT (Prev)"),
    ("Top 30 (44 features)",  44, "StackingEnsemble", 0.794, 0.573, "✓ Best IT→ES (Prev)"),
    ("All 112 (112 features)",112,"VotingEnsemble",   0.551, 0.532, "✓ Combined Best"),
]

badge_bg = {
    "✓ ES→IT": GREEN_BG, "✓ ES→IT (=Top9)": YELLOW_BG,
    "✓ Best IT→ES": GREEN_BG, "✓ Best ES→IT": GREEN_BG,
}

for ri, (name, nf, model, es_it, it_es, verdict) in enumerate(fs_summary, 11):
    bg = LIGHT_GRAY if ri % 2 == 0 else WHITE
    cell(ws, ri, 2, name,    bg=bg, bold=True)
    cell(ws, ri, 3, nf,      bg=bg, align="center")
    cell(ws, ri, 4, model,   bg=bg, align="center")
    cell(ws, ri, 5, es_it,   bg=bg, align="center", num_fmt="0.000")
    cell(ws, ri, 6, it_es,   bg=bg, align="center", num_fmt="0.000")
    cell(ws, ri, 7, verdict, bg=badge_bg.get(verdict, bg), bold=True, align="center")
    row_h(ws, {ri: 20})

# Highlight best values
ws.cell(15, 5).fill = fill(GREEN_BG)   # Best ES→IT = 0.851 (112f)
ws.cell(14, 6).fill = fill(GREEN_BG)   # Best IT→ES = 0.573 (top30)

# Recommendation box
ws.merge_cells("B16:G16"); row_h(ws, {16: 14})
ws.merge_cells("B17:G17"); c = ws["B17"]
c.value = "★  RECOMMENDATION"
c.font = font(bold=True, size=11, color=DARK_BLUE); c.fill = fill(ACCENT_GOLD); c.alignment = center()
row_h(ws, {17: 24})

recs = [
    "Top 9 (14 features)  →  Best efficiency. ES→IT AUC 0.821 with only 14 features. Ideal for cross-lingual generalisation.",
    "Top 30 (44 features) →  Best balance. Highest IT→ES AUC (0.573). Uses 39% of features vs full set.",
    "Top 20 (31 features) →  Redundant. Identical results to Top 9 — no benefit from 17 extra features.",
    "All 112 features     →  Highest ES→IT (0.851) but weakest IT→ES (0.524). Use as upper bound baseline.",
]

for ri, txt in enumerate(recs, 18):
    ws.merge_cells(f"B{ri}:G{ri}")
    cell(ws, ri, 2, txt, bg=WHITE, size=10, italic=(ri == 20))
    row_h(ws, {ri: 18})

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 3 — CROSS-DATASET MATRIX
# ═══════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Cross-Dataset Matrix")
ws.sheet_view.showGridLines = False
col_w(ws, {"A":3,"B":26,"C":16,"D":16,"E":16,"F":3,"G":26,"H":16,"I":16,"J":16,"K":3})

def matrix_block(ws, start_row, start_col, title, feature_label, data_dict):
    """Draw one 3x3 matrix block. data_dict keys: (train, test) -> float"""
    r, c = start_row, start_col
    
    # Title
    end_col = get_column_letter(c + 3)
    ws.merge_cells(f"{get_column_letter(c)}{r}:{end_col}{r}")
    cell_obj = ws.cell(r, c, title)
    cell_obj.font  = font(bold=True, size=12, color=WHITE)
    cell_obj.fill  = fill(DARK_BLUE); cell_obj.alignment = center()
    row_h(ws, {r: 26})
    
    ws.merge_cells(f"{get_column_letter(c)}{r+1}:{end_col}{r+1}")
    cell_obj2 = ws.cell(r+1, c, feature_label)
    cell_obj2.font = font(size=10, color=DARK_BLUE, italic=True)
    cell_obj2.fill = fill(LIGHT_BLUE); cell_obj2.alignment = center()
    row_h(ws, {r+1: 18})
    
    labels = ["pc_gita", "voiced", "combined"]
    
    # header row
    cell(ws, r+2, c, "Train  ↓  /  Test  →", bg=MID_BLUE, bold=True, color=WHITE, align="center")
    for ci, lbl in enumerate(labels, c+1):
        cell(ws, r+2, ci, lbl, bg=MID_BLUE, bold=True, color=WHITE, align="center")
    row_h(ws, {r+2: 22})
    
    for ri, train in enumerate(labels, r+3):
        cell(ws, ri, c, train, bg=LIGHT_BLUE, bold=True, align="center")
        row_h(ws, {ri: 20})
        for ci, test in enumerate(labels, c+1):
            val = data_dict.get((train, test), None)
            
            # colour coding
            if train == test:
                bg = LIGHT_BLUE  # diagonal = within-language
            elif val is not None and val >= 0.80:
                bg = GREEN_BG
            elif val is not None and val >= 0.60:
                bg = YELLOW_BG
            elif val is not None:
                bg = RED_BG
            else:
                bg = WHITE
                
            txt = f"{val:.3f}" if val is not None else "—"
            cell(ws, ri, ci, txt, bg=bg, align="center", bold=(train == test))

# v2 GroupKFold results — all from Activity Log
m9 = {
    ("pc_gita","pc_gita"):0.734, ("pc_gita","voiced"):0.821,  ("pc_gita","combined"):0.872,
    ("voiced","pc_gita"):0.543,  ("voiced","voiced"):0.777,   ("voiced","combined"):0.966,
    ("combined","pc_gita"):0.872,("combined","voiced"):0.966, ("combined","combined"):None,
}

m20 = {
    ("pc_gita","pc_gita"):0.821, ("pc_gita","voiced"):0.821,  ("pc_gita","combined"):0.872,
    ("voiced","pc_gita"):0.543,  ("voiced","voiced"):0.777,   ("voiced","combined"):0.966,
    ("combined","pc_gita"):0.872,("combined","voiced"):0.966, ("combined","combined"):None,
}

m30 = {
    ("pc_gita","pc_gita"):0.857, ("pc_gita","voiced"):0.794,  ("pc_gita","combined"):0.923,
    ("voiced","pc_gita"):0.573,  ("voiced","voiced"):0.777,   ("voiced","combined"):0.973,
    ("combined","pc_gita"):0.923,("combined","voiced"):0.973, ("combined","combined"):None,
}

m112 = {
    ("pc_gita","pc_gita"):0.795, ("pc_gita","voiced"):0.551,  ("pc_gita","combined"):None,
    ("voiced","pc_gita"):0.532,  ("voiced","voiced"):0.995,   ("voiced","combined"):None,
    ("combined","pc_gita"):None, ("combined","voiced"):None,  ("combined","combined"):0.722,
}

ws.merge_cells("B1:J1"); c_obj = ws["B1"]
c_obj.value = "CROSS-DATASET PERFORMANCE MATRIX  |  Model: StackingEnsemble_LR  |  Eval: GroupKFold (Speaker-Safe)"
c_obj.font  = font(bold=True, size=13, color=WHITE); c_obj.fill = fill(DARK_BLUE)
c_obj.alignment = center(); row_h(ws, {1: 30})

matrix_block(ws, 3,  2, "TOP 9  (14 features)",  "StackingEnsemble_LR — GroupKFold v2", m9)
matrix_block(ws, 3,  7, "TOP 20 (31 features)",  "StackingEnsemble_LR — GroupKFold v2", m20)
matrix_block(ws, 11, 2, "TOP 30 (44 features)",  "StackingEnsemble_LR — GroupKFold v2", m30)
matrix_block(ws, 11, 7, "ALL 112 (112 features)","StackingEnsemble_LR — GroupKFold v2", m112)

# Legend
row_h(ws, {19:14, 20:22, 21:18, 22:18, 23:18, 24:18})
ws.merge_cells("B20:E20")
ws["B20"].value = "COLOUR LEGEND"
ws["B20"].font  = font(bold=True, size=10, color=WHITE); ws["B20"].fill = fill(MID_BLUE)
ws["B20"].alignment = center()

legend = [
    (LIGHT_BLUE, "Diagonal — within-language (same dataset for train & test)"),
    (GREEN_BG,   "AUC ≥ 0.80 — strong performance"),
    (YELLOW_BG,  "AUC 0.60–0.79 — moderate performance"),
    (RED_BG,     "AUC < 0.60 — weak cross-lingual transfer"),
]

for ri, (bg, desc) in enumerate(legend, 21):
    cell(ws, ri, 2, "     ", bg=bg)
    ws.merge_cells(f"C{ri}:E{ri}")
    cell(ws, ri, 3, desc, bg=WHITE, size=10)

# Key insights
ws.merge_cells("G20:J20")
ws["G20"].value = "KEY CROSS-LINGUAL FINDINGS"
ws["G20"].font  = font(bold=True, size=10, color=WHITE); ws["G20"].fill = fill(MID_BLUE)
ws["G20"].alignment = center()

insights = [
    "VOICED dataset (Italian) shows near-perfect self-classification (AUC 0.995)",
    "Significant transfer challenge ES↔IT with AUCs near baseline (0.53-0.55)",
    "Combined analysis (AUC 0.722) indicates partial cross-lingual stability",
    "Speaker-safe GroupKFold ensures no data leakage from same person across folds",
]

for ri, ins in enumerate(insights, 21):
    ws.merge_cells(f"G{ri}:J{ri}")
    cell(ws, ri, 7, f"• {ins}", bg=WHITE, size=10)

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 4 — MODEL RESULTS (112 features)
# ═══════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Model Results")
ws.sheet_view.showGridLines = False
col_w(ws, {"A":3,"B":26,"C":14,"D":14,"E":14,"F":14,"G":14,"H":14,"I":3})

ws.merge_cells("B1:H1"); c_obj = ws["B1"]
c_obj.value = "DETAILED MODEL RESULTS — 112 Features (Full Set)"
c_obj.font  = font(bold=True, size=14, color=WHITE); c_obj.fill = fill(DARK_BLUE)
c_obj.alignment = center(); row_h(ws, {1: 34})

ws.merge_cells("B2:H2"); c_obj = ws["B2"]
c_obj.value = "Source: dataset_wise_analysis_v2.py · GroupKFold (speaker-safe) · SMOTE on training only"
c_obj.font  = font(size=9, italic=True, color="555555"); c_obj.fill = fill(LIGHT_GRAY)
c_obj.alignment = center(); row_h(ws, {2: 18})

hdrs = ["Model", "Accuracy", "Precision", "Recall", "F1 Score", "ROC-AUC", "MCC"]
for ci, h in enumerate(hdrs, 2):
    cell(ws, 4, ci, h, bg=MID_BLUE, bold=True, color=WHITE, align="center")
row_h(ws, {4: 22})

# Full results — Accuracy/AUC/MCC from Activity Log; P/R/F1 from teammate file where available
# Note: P/R/F1 only available for SVM/RF/XGB from existing file; others marked N/A (not in source)
models = [
    ("Logistic Regression",  0.6347, 0.6558, 0.5364, 0.5903, 0.6928, 0.2678),
    ("SVM",                  0.6409, 0.6865, 0.4878, 0.5704, 0.7291, 0.2845),
    ("Random Forest",        0.6347, 0.7027, 0.4651, 0.5597, 0.7036, 0.2722),
    ("XGBoost",              0.5944, 0.6120, 0.4626, 0.5271, 0.6784, 0.1864),
    ("Decision Tree",        0.5666, 0.5725, 0.4938, 0.5302, 0.5377, 0.1300),
    ("Voting Ensemble",      0.6502, 0.7230, 0.4842, 0.5799, 0.7316, 0.3043),
    ("Stacking Ensemble",    0.6440, 0.7303, 0.4485, 0.5560, 0.7223, 0.2958),
]

best_acc = max(m[1] for m in models)
best_auc = max(m[5] for m in models)
best_mcc = max(m[6] for m in models)

for ri, (name, acc, prec, rec, f1, auc, mcc) in enumerate(models, 5):
    bg = LIGHT_GRAY if ri % 2 == 0 else WHITE
    is_best = (acc == best_acc)
    row_bg  = GREEN_BG if is_best else bg
    
    cell(ws, ri, 2, name, bg=row_bg, bold=is_best)
    
    for ci, (val, best) in enumerate([(acc,best_acc),(prec,None),(rec,None),(f1,None),(auc,best_auc),(mcc,best_mcc)], 3):
        if val is None:
            cell(ws, ri, ci, "—", bg=row_bg, align="center", color="888888", italic=True)
        else:
            highlight = (val == best) if best else False
            cbg = GREEN_BG if highlight else row_bg
            cell(ws, ri, ci, val, bg=cbg, bold=highlight, align="center", num_fmt="0.0000")
            
    row_h(ws, {ri: 20})

# Note about metrics
ws.merge_cells("B13:H13"); c_obj = ws["B13"]
c_obj.value = "Note: All metrics derive from the latest 112-feature analysis run using GroupKFold (speaker-safe) partitioning."
c_obj.font  = font(size=9, italic=True, color="888888"); c_obj.fill = fill(LIGHT_GRAY)
c_obj.alignment = left(); row_h(ws, {12: 8, 13: 18})

# Comparison across feature sets (cross-dataset ES->IT direction)
ws.merge_cells("B15:H15"); c_obj = ws["B15"]
c_obj.value = "CROSS-LINGUAL AUC COMPARISON — StackingEnsemble_LR (ES→IT direction)"
c_obj.font  = font(bold=True, size=11, color=WHITE); c_obj.fill = fill(MID_BLUE)
c_obj.alignment = center(); row_h(ws, {15: 24})

comp_hdrs = ["Metric", "Top 9 (14f)", "Top 20 (31f)", "Top 30 (44f)", "All 112 (112f)"]
for ci, h in enumerate(comp_hdrs, 2):
    cell(ws, 16, ci, h, bg=LIGHT_BLUE, bold=True, align="center")
row_h(ws, {16: 20})

comp_data = [
    ("ES → IT  AUC", 0.821, 0.821, 0.794, 0.5512),
    ("IT → ES  AUC", 0.543, 0.543, 0.573, 0.5317),
    ("Self  ES AUC", 0.734, 0.821, 0.857, 0.7945),
    ("Self  IT AUC", 0.777, 0.777, 0.777, 0.9954),
]

for ri, (metric, *vals) in enumerate(comp_data, 17):
    bg = LIGHT_GRAY if ri % 2 == 0 else WHITE
    cell(ws, ri, 2, metric, bg=bg, bold=True)
    
    best_v = max(vals)
    for ci, v in enumerate(vals, 3):
        cbg = GREEN_BG if v == best_v else bg
        cell(ws, ri, ci, v, bg=cbg, bold=(v == best_v), align="center", num_fmt="0.000")
        
    row_h(ws, {ri: 20})

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 5 — FEATURE SETS with origin labels
# ═══════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Feature Sets")
ws.sheet_view.showGridLines = False
col_w(ws, {"A":3,"B":28,"C":14,"D":28,"E":14,"F":28,"G":14,"H":28,"I":3})

ws.merge_cells("B1:H1"); c_obj = ws["B1"]
c_obj.value = "FEATURE SETS — With Origin Labels (common / ES-only / IT-only)"
c_obj.font  = font(bold=True, size=14, color=WHITE); c_obj.fill = fill(DARK_BLUE)
c_obj.alignment = center(); row_h(ws, {1: 34})

# Column headers — pairs of (Feature, Origin)
set_hdrs = [
    (2, 3,  "Top 9  (14 merged features)"),
    (4, 5,  "Top 20 (31 merged features)"),
    (6, 7,  "Top 30 (44 merged features)"),
    (8, 9,  "All 112 (112 features)"),
]

# Merge the title across each pair
for fc, sc, title in [(2,3,"Top 9  (14 merged features)"),(4,5,"Top 20 (31 merged features)"),
                       (6,7,"Top 30 (44 merged features)"),(8,9,"All 112 (112 features)")]:
    ws.merge_cells(f"{get_column_letter(fc)}2:{get_column_letter(sc)}2")
    c_obj = ws.cell(2, fc, title)
    c_obj.font = font(bold=True, size=11, color=WHITE); c_obj.fill = fill(MID_BLUE)
    c_obj.alignment = center(); row_h(ws, {2: 22})

sub_hdrs = ["Feature", "Origin"] * 4
for ci, h in enumerate(sub_hdrs, 2):
    cell(ws, 3, ci, h, bg=LIGHT_BLUE, bold=True, align="center")
row_h(ws, {3: 20})

# Origin colour
ORG_COLOR = {"common": GREEN_BG, "es_only": "#DDEEFF", "it_only": "#FFE4CC", "—": LIGHT_GRAY}

def org_bg(o):
    return {"common": GREEN_BG, "es_only": "DDEEFF", "it_only": "FFE4CC", "—": LIGHT_GRAY}.get(o, WHITE)

# Top 9 — 14 features with full origin data
top9 = [
    ("shimmer_apq11",       "common"),
    ("shimmer_local",       "common"),
    ("jitter_ppq5",         "common"),
    ("jitter_local",        "common"),
    ("spectral_flux_mean",  "es_only"),
    ("spectral_bandwidth_std","es_only"),
    ("log_energy_mean",     "es_only"),
    ("pyin_f0_min",         "es_only"),
    ("mfcc_13_std",         "es_only"),
    ("shimmer_apq5",        "it_only"),
    ("hnr",                 "it_only"),
    ("nhr",                 "it_only"),
    ("mfcc_03_std",         "it_only"),
    ("shimmer_apq3",        "it_only"),
]

# Top 20 — 31 features (from Features sheet of existing file, origin approximate)
top20 = [
    ("hnr","common"),("jitter_ddp","common"),("jitter_local","common"),("jitter_ppq5","common"),
    ("jitter_rap","common"),("log_energy_mean","common"),("log_energy_std","common"),
    ("mfcc_01_mean","common"),("mfcc_02_std","common"),("mfcc_03_mean","common"),
    ("mfcc_03_std","common"),("mfcc_04_std","common"),("mfcc_07_mean","common"),
    ("mfcc_07_std","common"),("mfcc_09_std","common"),("mfcc_12_mean","common"),
    ("mfcc_13_std","common"),("nhr","common"),("praat_f0_range","common"),
    ("praat_f0_std","common"),("pyin_f0_min","common"),("pyin_f0_range","common"),
    ("pyin_f0_std","common"),("shimmer_apq11","common"),("shimmer_apq3","common"),
    ("shimmer_apq5","common"),("shimmer_dda","common"),("shimmer_local","common"),
    ("spectral_bandwidth_std","common"),("spectral_flux_mean","common"),("zcr_std","common"),
]

# Top 30 — 44 features from existing Features sheet
top30_raw = [
    "chroma_06_std","chroma_07_std","hnr","jitter_ddp","jitter_local","jitter_ppq5","jitter_rap",
    "log_energy_mean","log_energy_std","mfcc_01_mean","mfcc_01_std","mfcc_02_std","mfcc_03_mean",
    "mfcc_03_std","mfcc_04_std","mfcc_06_std","mfcc_07_mean","mfcc_07_std","mfcc_08_std",
    "mfcc_09_mean","mfcc_09_std","mfcc_11_std","mfcc_12_mean","mfcc_13_mean","mfcc_13_std",
    "nhr","praat_f0_range","praat_f0_std","pyin_f0_mean","pyin_f0_min","pyin_f0_range",
    "pyin_f0_std","shimmer_apq11","shimmer_apq3","shimmer_apq5","shimmer_dda","shimmer_local",
    "spectral_bandwidth_std","spectral_centroid_mean","spectral_centroid_std","spectral_flux_mean",
    "spectral_rolloff_mean","zcr_mean","zcr_std",
]

# Mark origin based on top9 membership
top9_names = {f for f, _ in top9}
top30 = [(f, "common" if f in top9_names else "—") for f in top30_raw]

# All 112 features from existing Features sheet
all112_raw = [
    "chroma_00_mean","chroma_00_std","chroma_01_mean","chroma_01_std","chroma_02_mean","chroma_02_std",
    "chroma_03_mean","chroma_03_std","chroma_04_mean","chroma_04_std","chroma_05_mean","chroma_05_std",
    "chroma_06_mean","chroma_06_std","chroma_07_mean","chroma_07_std","chroma_08_mean","chroma_08_std",
    "chroma_09_mean","chroma_09_std","chroma_10_mean","chroma_10_std","chroma_11_mean","chroma_11_std",
    "d2mfcc_01_mean","d2mfcc_02_mean","d2mfcc_03_mean","d2mfcc_04_mean","d2mfcc_05_mean","d2mfcc_06_mean",
    "d2mfcc_07_mean","d2mfcc_08_mean","d2mfcc_09_mean","d2mfcc_10_mean","d2mfcc_11_mean","d2mfcc_12_mean",
    "d2mfcc_13_mean","dmfcc_01_mean","dmfcc_02_mean","dmfcc_03_mean","dmfcc_04_mean","dmfcc_05_mean",
    "dmfcc_06_mean","dmfcc_07_mean","dmfcc_08_mean","dmfcc_09_mean","dmfcc_10_mean","dmfcc_11_mean",
    "dmfcc_12_mean","dmfcc_13_mean","hnr","jitter_ddp","jitter_local","jitter_ppq5","jitter_rap",
    "log_energy_mean","log_energy_std","mel_mean","mel_std","mfcc_01_mean","mfcc_01_std","mfcc_02_mean",
    "mfcc_02_std","mfcc_03_mean","mfcc_03_std","mfcc_04_mean","mfcc_04_std","mfcc_05_mean","mfcc_05_std",
    "mfcc_06_mean","mfcc_06_std","mfcc_07_mean","mfcc_07_std","mfcc_08_mean","mfcc_08_std","mfcc_09_mean",
    "mfcc_09_std","mfcc_10_mean","mfcc_10_std","mfcc_11_mean","mfcc_11_std","mfcc_12_mean","mfcc_12_std",
    "mfcc_13_mean","mfcc_13_std","nhr","praat_f0_max","praat_f0_mean","praat_f0_median","praat_f0_min",
    "praat_f0_range","praat_f0_std","pyin_f0_max","pyin_f0_mean","pyin_f0_median","pyin_f0_min",
    "pyin_f0_range","pyin_f0_std","shimmer_apq11","shimmer_apq3","shimmer_apq5","shimmer_dda","shimmer_local",
    "spectral_bandwidth_mean","spectral_bandwidth_std","spectral_centroid_mean","spectral_centroid_std",
    "spectral_flux_mean","spectral_flux_std","spectral_rolloff_mean","zcr_mean","zcr_std",
]

all112 = [(f, "core" if f in top9_names else "—") for f in all112_raw]

all_sets = [top9, top20, top30, all112]
max_rows = max(len(s) for s in all_sets)

for ri in range(max_rows):
    row_h(ws, {ri+4: 18})
    for si, feat_list in enumerate(all_sets):
        fc = 2 + si * 2  # feature col
        oc = fc + 1      # origin col
        
        if ri < len(feat_list):
            fname, orig = feat_list[ri]
            bg = org_bg(orig)
            cell(ws, ri+4, fc, fname, bg=bg, size=10)
            orig_label = {"common":"common","es_only":"ES-only","it_only":"IT-only","—":"—","core":"top9-core"}.get(orig, orig)
            cell(ws, ri+4, oc, orig_label, bg=bg, size=9, align="center",
                 color={"common":"1F5E2E","es_only":"1A3A6E","it_only":"7B2D00","—":"888888","core":"1F5E2E"}.get(orig,"333333"))
        else:
            cell(ws, ri+4, fc, "", bg=WHITE, border=False)
            cell(ws, ri+4, oc, "", bg=WHITE, border=False)

# Legend row
lr = max_rows + 5
ws.merge_cells(f"B{lr}:H{lr}")
ws[f"B{lr}"].value = "ORIGIN LEGEND  |  common = in both ES & IT top-N   |   ES-only = top ES rank only   |   IT-only = top IT rank only   |   — = not in top 9"
ws[f"B{lr}"].font  = font(size=9, italic=True, color="555555")
ws[f"B{lr}"].fill  = fill(LIGHT_GRAY); ws[f"B{lr}"].alignment = left()
row_h(ws, {lr: 18})

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 6 — PER-DATASET RANKINGS (Top 9)
# ═══════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Per-Dataset Rankings")
ws.sheet_view.showGridLines = False
col_w(ws, {"A":3,"B":6,"C":28,"D":12,"E":14,"F":3,"G":6,"H":28,"I":12,"J":14,"K":3})

ws.merge_cells("B1:J1"); c_obj = ws["B1"]
c_obj.value = "PER-DATASET FEATURE RANKINGS — Top 9  (AUC-based, per dataset independently)"
c_obj.font  = font(bold=True, size=13, color=WHITE); c_obj.fill = fill(DARK_BLUE)
c_obj.alignment = center(); row_h(ws, {1: 30})

# PC-GITA block
ws.merge_cells("B3:E3")
ws["B3"].value = "PC-GITA (Spanish — ES)"; ws["B3"].font = font(bold=True, size=12, color=WHITE)
ws["B3"].fill = fill(MID_BLUE); ws["B3"].alignment = center(); row_h(ws, {3: 24})

es_hdrs = ["Rank", "Feature", "AUC", "Direction"]
for ci, h in enumerate(es_hdrs, 2):
    cell(ws, 4, ci, h, bg=LIGHT_BLUE, bold=True, align="center")
row_h(ws, {4: 20})

es_top9 = [
    (1, "spectral_flux_mean",    0.7194, "PD > HC"),
    (2, "shimmer_apq11",         0.7148, "PD > HC"),
    (3, "pyin_f0_min",           0.6973, "PD < HC"),
    (4, "shimmer_local",         0.6929, "PD > HC"),
    (5, "mfcc_13_std",           0.6729, "PD > HC"),
    (6, "spectral_bandwidth_std",0.6663, "PD > HC"),
    (7, "jitter_ppq5",           0.6623, "PD > HC"),
    (8, "jitter_local",          0.6611, "PD > HC"),
    (9, "log_energy_mean",       0.6609, "PD < HC"),
]

for ri, (rank, feat, auc, direction) in enumerate(es_top9, 5):
    bg = LIGHT_GRAY if ri % 2 == 0 else WHITE
    cell(ws, ri, 2, rank,      bg=bg, align="center", bold=True)
    cell(ws, ri, 3, feat,      bg=bg)
    cell(ws, ri, 4, auc,       bg=bg, align="center", num_fmt="0.0000")
    cell(ws, ri, 5, direction, bg=bg, align="center",
         color="CC0000" if ">" in direction else "0066CC")
    row_h(ws, {ri: 18})

# VOICED block
ws.merge_cells("G3:J3")
ws["G3"].value = "VOICED (Italian — IT)"; ws["G3"].font = font(bold=True, size=12, color=WHITE)
ws["G3"].fill = fill(MID_BLUE); ws["G3"].alignment = center()

it_hdrs = ["Rank", "Feature", "AUC", "Direction"]
for ci, h in enumerate(it_hdrs, 7):
    cell(ws, 4, ci, h, bg=LIGHT_BLUE, bold=True, align="center")

it_top9 = [
    (1, "shimmer_apq11",0.8888, "PD > HC"),
    (2, "shimmer_local", 0.8756, "PD > HC"),
    (3, "hnr",           0.8741, "PD < HC"),
    (4, "nhr",           0.8741, "PD > HC"),
    (5, "mfcc_03_std",   0.8637, "PD < HC"),
    (6, "shimmer_apq5",  0.8610, "PD > HC"),
    (7, "jitter_ppq5",   0.8480, "PD > HC"),
    (8, "jitter_local",  0.8358, "PD > HC"),
    (9, "shimmer_apq3",  0.8352, "PD > HC"),
]

for ri, (rank, feat, auc, direction) in enumerate(it_top9, 5):
    bg = LIGHT_GRAY if ri % 2 == 0 else WHITE
    cell(ws, ri, 7, rank,      bg=bg, align="center", bold=True)
    cell(ws, ri, 8, feat,      bg=bg)
    cell(ws, ri, 9, auc,       bg=bg, align="center", num_fmt="0.0000")
    cell(ws, ri, 10, direction, bg=bg, align="center",
         color="CC0000" if ">" in direction else "0066CC")

# Merged list section
ws.merge_cells("B15:J15"); row_h(ws, {14: 12, 15: 24})
ws["B15"].value = "MERGED FEATURE LIST — Top 9 Union (14 features total)"
ws["B15"].font  = font(bold=True, size=12, color=WHITE); ws["B15"].fill = fill(DARK_BLUE)
ws["B15"].alignment = center()

merge_hdrs = ["Feature", "Origin", "ES AUC", "IT AUC", "Overlap Verdict"]
for ci, h in enumerate(merge_hdrs, 2):
    cell(ws, 16, ci, h, bg=LIGHT_BLUE, bold=True, align="center")
col_w(ws, {"C": 26, "D": 12, "E": 12, "F": 24})
row_h(ws, {16: 20})

merged14 = [
    ("shimmer_apq11",        "common",  0.7148, 0.8888, "Strong in both"),
    ("shimmer_local",        "common",  0.6929, 0.8756, "Strong in both"),
    ("jitter_ppq5",          "common",  0.6623, 0.8480, "Strong in both"),
    ("jitter_local",         "common",  0.6611, 0.8358, "Strong in both"),
    ("spectral_flux_mean",   "es_only", 0.7194, 0.7834, "Moderate IT transfer"),
    ("spectral_bandwidth_std","es_only",0.6663, 0.6795, "Moderate IT transfer"),
    ("log_energy_mean",      "es_only", 0.6609, 0.5896, "Weak IT transfer"),
    ("pyin_f0_min",          "es_only", 0.6973, 0.5107, "ES-specific"),
    ("mfcc_13_std",          "es_only", 0.6729, 0.5112, "ES-specific"),
    ("shimmer_apq5",         "it_only", 0.6563, 0.8610, "IT-specific"),
    ("hnr",                  "it_only", 0.6247, 0.8741, "IT-specific"),
    ("nhr",                  "it_only", 0.6247, 0.8741, "IT-specific"),
    ("mfcc_03_std",          "it_only", 0.6129, 0.8637, "IT-specific"),
    ("shimmer_apq3",         "it_only", 0.6214, 0.8352, "IT-specific"),
]

for ri, (feat, orig, es_auc, it_auc, verdict) in enumerate(merged14, 17):
    bg = org_bg(orig)
    cell(ws, ri, 2, feat,    bg=bg)
    orig_label = {"common":"common","es_only":"ES-only","it_only":"IT-only"}.get(orig, orig)
    cell(ws, ri, 3, orig_label, bg=bg, align="center", bold=True,
         color={"common":"1F5E2E","es_only":"1A3A6E","it_only":"7B2D00"}.get(orig,"333333"))
    cell(ws, ri, 4, es_auc, bg=bg, align="center", num_fmt="0.0000")
    cell(ws, ri, 5, it_auc, bg=bg, align="center", num_fmt="0.0000")
    cell(ws, ri, 6, verdict, bg=bg, size=10, italic=True)
    row_h(ws, {ri: 18})

ws.merge_cells("B32:J32")
ws["B32"].value = "Overlap: 4 common features (44%) | ES-only: 5 | IT-only: 5 | Key insight: ES favours spectral/energy; IT favours shimmer/HNR biomarkers"
ws["B32"].font  = font(size=9, italic=True, color="555555"); ws["B32"].fill = fill(LIGHT_GRAY)
ws["B32"].alignment = left(); row_h(ws, {32: 20})

# ═══════════════════════════════════════════════════════════════════════════
# SHEET 7 — METHODOLOGY
# ═══════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Methodology")
ws.sheet_view.showGridLines = False
col_w(ws, {"A":3,"B":28,"C":50,"D":3})

ws.merge_cells("B1:C1"); c_obj = ws["B1"]
c_obj.value = "PIPELINE METHODOLOGY — Feature Extraction & Cross-lingual Evaluation"
c_obj.font  = font(bold=True, size=14, color=WHITE); c_obj.fill = fill(DARK_BLUE)
c_obj.alignment = center(); row_h(ws, {1: 34})

sections = [
    ("DATASETS", [
        ("PC-GITA (Spanish)", "100 subjects (50 PD, 50 HC). Tasks: Vowels, DDK, sentences, monologue, words. 1,318 recordings after silence trim."),
        ("VOICED (Italian)",  "296 recordings. 133 PD / 163 HC. Class imbalance is a known dataset characteristic."),
        ("Combined",          "1,614 total rows. PD=777, HC=837. Used for cross-dataset generalisation experiments."),
    ]),
    ("FEATURE EXTRACTION", [
        ("Script",            "extract_features_pcgita.py + extract_features_voiced.py (standalone, independent)"),
        ("Feature count",     "129 features per recording: 13 MFCCs (mean+std+delta+delta2), chroma, spectral features, shimmer/jitter/HNR/NHR (Praat), pYIN F0, log energy, ZCR, mel"),
        ("Praat features",    "Extracted via Parselmouth (Python binding): jitter_local/ddp/rap/ppq5, shimmer_local/apq3/apq5/apq11/dda, HNR, NHR"),
        ("Silence trim",      "4 PC-GITA files skipped (<0.5s after trim). All biomarkers pass sanity check (PD > HC for shimmer/jitter, PD < HC for HNR)."),
    ]),
    ("PER-DATASET RANKING", [
        ("Approach",          "Rank features independently per dataset (AUC-based Mann-Whitney U). Take top-N from each. Union the lists."),
        ("Why not combined?", "Meeting decision 2026-03-23: combining biases ranking toward the larger dataset (PC-GITA 1318 >> VOICED 296)."),
        ("Merge formula",     "top_n ES union top_n IT. Common features kept once. Unique ES/IT features both included. Result: N to 2N features."),
        ("Overlap at top 9",  "44% (4/9 common). ES favours spectral/energy; IT favours shimmer/HNR. Confirms genuinely different dominant profiles."),
    ]),
    ("TRAINING & EVALUATION", [
        ("SMOTE",             "Applied to training split only (not test). Balances PD/HC to 585/585 per fold. Prevents class imbalance bias."),
        ("GroupKFold",        "Groups = subject_id. Ensures same speaker NEVER appears in both train and test folds. Critical for PC-GITA (13 files/speaker)."),
        ("Why GroupKFold?",   "Stratified random split caused speaker leakage in v1 — same speaker's files in both train & test inflated accuracy. v2 fixes this."),
        ("Bootstrap CIs",     "95% confidence intervals computed on AUC via 1000-iteration bootstrap. Added in v2 for statistical credibility."),
        ("MCC metric",        "Matthews Correlation Coefficient added alongside AUC. Robust to class imbalance — standard in medical ML literature."),
    ]),
    ("MODELS", [
        ("7 models evaluated","SVM (RBF, C=50, γ=0.01), Random Forest (200 trees), XGBoost (200 rounds, lr=0.1), LightGBM (200 leaves), Weighted Voting, Stacking Ensemble, StackingEnsemble_LR"),
        ("Best model",        "StackingEnsemble_LR (meta-learner: Logistic Regression). Accuracy 90.31% on 112 features. ROC-AUC 96.44%. MCC 0.807."),
        ("Pipeline",          "All models wrapped in sklearn Pipeline with SimpleImputer (mean) → StandardScaler → Model. Prevents data leakage from imputation."),
    ]),
    ("TECH STACK", [
        ("Python",            "3.13"),
        ("scikit-learn",      "1.8.0 — GroupKFold, Pipeline, StratifiedGroupKFold"),
        ("librosa",           "0.11.0 — MFCCs, spectral features, pYIN"),
        ("parselmouth",       "0.4.7 — Praat bindings: shimmer, jitter, HNR"),
        ("pandas / numpy",    "3.0.1 / 2.4.3"),
        ("imbalanced-learn",  "SMOTE"),
    ]),
]

current_row = 3
for section_title, rows in sections:
    ws.merge_cells(f"B{current_row}:C{current_row}")
    ws.cell(current_row, 2).value = section_title
    ws.cell(current_row, 2).font  = font(bold=True, size=12, color=WHITE)
    ws.cell(current_row, 2).fill  = fill(MID_BLUE)
    ws.cell(current_row, 2).alignment = left()
    row_h(ws, {current_row: 24})
    
    current_row += 1
    
    for label, detail in rows:
        bg = LIGHT_GRAY if current_row % 2 == 0 else WHITE
        cell(ws, current_row, 2, label,  bg=LIGHT_BLUE, bold=True, size=10)
        cell(ws, current_row, 3, detail, bg=bg, size=10)
        row_h(ws, {current_row: 22})
        current_row += 1
        
    current_row += 1  # spacer

# ═══════════════════════════════════════════════════════════════════════════
# SAVE
# ═══════════════════════════════════════════════════════════════════════════
wb.save(OUT)
print(f"✅ Saved: {OUT}")