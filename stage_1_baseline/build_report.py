"""
Build a comprehensive multi-sheet Excel report from PD speech classification results.
Sheets: Raw Data, Key Metrics, Cross-Dataset, Feature Rankings, Analytical Summary
"""

import os, json, glob
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

RESULTS_DIR = os.path.join(os.path.dirname(__file__), "results")
OUTPUT_FILE = os.path.join(os.path.dirname(__file__), "PD_Speech_Analysis_Report.xlsx")

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SUBHEADER_FILL = PatternFill("solid", fgColor="D6E4F0")
SUBHEADER_FONT = Font(bold=True, color="1F4E79", size=10)
BEST_FILL = PatternFill("solid", fgColor="C6EFCE")
WORST_FILL = PatternFill("solid", fgColor="FFC7CE")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
NUM_FMT_PCT = "0.00%"
NUM_FMT_DEC = "0.0000"


def parse_experiment_name(dirname):
    """Extract feature_set, sample_rate, and experiment_type from directory name."""
    parts = dirname.lower()
    # Determine type
    if parts.startswith("evaluation_features_"):
        exp_type = "Feature Evaluation"
        rest = parts.replace("evaluation_features_", "")
    elif parts.startswith("evaluation_training_"):
        exp_type = "Training Evaluation"
        rest = parts.replace("evaluation_training_", "")
    else:
        return None

    # Extract sample rate
    sample_rate = "N/A"
    for sr in ["16k", "10k", "8k"]:
        if f"_{sr}_" in rest:
            sample_rate = sr
            break

    # Extract feature set
    feature_set = rest.split("_20")[0]
    if sample_rate != "N/A":
        feature_set = feature_set.replace(f"_{sample_rate}", "")

    return {
        "experiment": dirname,
        "type": exp_type,
        "feature_set": feature_set,
        "sample_rate": sample_rate,
    }


def load_all_csv_results():
    """Load all analysis_summary.csv files into a single DataFrame."""
    rows = []
    eval_dirs = sorted(glob.glob(os.path.join(RESULTS_DIR, "evaluation_*")))
    for d in eval_dirs:
        csv_path = os.path.join(d, "csv_results", "analysis_summary.csv")
        if not os.path.exists(csv_path):
            continue
        info = parse_experiment_name(os.path.basename(d))
        if info is None:
            continue
        df = pd.read_csv(csv_path)
        df["Experiment"] = info["experiment"]
        df["Type"] = info["type"]
        df["Feature_Set"] = info["feature_set"]
        df["Sample_Rate"] = info["sample_rate"]
        rows.append(df)
    if not rows:
        return pd.DataFrame()
    return pd.concat(rows, ignore_index=True)


def load_ranking_data():
    """Load final ranking CSVs."""
    final_dir = os.path.join(RESULTS_DIR, "final")
    rankings = {}
    if not os.path.isdir(final_dir):
        return rankings
    for d in sorted(os.listdir(final_dir)):
        full = os.path.join(final_dir, d)
        if not os.path.isdir(full):
            continue
        for f in os.listdir(full):
            if f.endswith(".csv"):
                key = f"{d}/{f}"
                rankings[key] = pd.read_csv(os.path.join(full, f))
    return rankings


def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def style_subheader_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER


def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value) if cell.value else ""
                max_len = max(max_len, len(val))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 3, 40)


def apply_borders(ws, min_row, max_row, max_col):
    for r in range(min_row, max_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).border = THIN_BORDER


def write_raw_data(writer, df):
    """Sheet 1: All raw results."""
    cols = ["Experiment", "Type", "Feature_Set", "Sample_Rate", "Analysis",
            "Model", "Accuracy", "F1", "AUC", "AUC_CI_lo", "AUC_CI_hi", "MCC"]
    out = df[cols].copy()
    out.to_excel(writer, sheet_name="Raw Data", index=False)


def write_key_metrics(writer, df):
    """Sheet 2: Best model per experiment+analysis, pivoted."""
    best = df.loc[df.groupby(["Experiment", "Analysis"])["AUC"].idxmax()]
    cols = ["Feature_Set", "Sample_Rate", "Analysis", "Model",
            "Accuracy", "F1", "AUC", "MCC"]
    best[cols].to_excel(writer, sheet_name="Key Metrics", index=False)


def write_cross_dataset(writer, df):
    """Sheet 3: Cross-dataset generalization analysis."""
    cross = df[df["Analysis"].str.contains("→", na=False)].copy()
    if cross.empty:
        cross = df[df["Analysis"].str.contains("→", na=False)].copy()
    if cross.empty:
        # Try arrow character
        cross = df[~df["Analysis"].str.contains("Self|Combined", na=False)].copy()

    if not cross.empty:
        pivot = cross.groupby(["Feature_Set", "Sample_Rate", "Analysis", "Model"]).agg(
            Accuracy=("Accuracy", "mean"),
            F1=("F1", "mean"),
            AUC=("AUC", "mean"),
            MCC=("MCC", "mean"),
        ).reset_index()
        pivot.to_excel(writer, sheet_name="Cross-Dataset", index=False)
    else:
        pd.DataFrame({"Note": ["No cross-dataset results found"]}).to_excel(
            writer, sheet_name="Cross-Dataset", index=False)


def write_feature_rankings(writer, rankings):
    """Sheet 4+: Feature ranking data."""
    for key, rdf in rankings.items():
        # Create readable sheet name (max 31 chars)
        name = key.split("/")[-1].replace(".csv", "")
        name = name[:31]
        rdf.to_excel(writer, sheet_name=name, index=False)


def write_analytical_summary(writer, df):
    """Sheet: Analytical Summary with aggregated insights."""
    rows = []

    # 1. Best overall model per feature set (self-evaluation only)
    self_df = df[df["Analysis"].str.contains("Self", na=False)]
    if not self_df.empty:
        for fs in self_df["Feature_Set"].unique():
            subset = self_df[self_df["Feature_Set"] == fs]
            best_idx = subset["AUC"].idxmax()
            best = subset.loc[best_idx]
            rows.append({
                "Category": "Best Self-Eval Model",
                "Feature Set": fs,
                "Detail": f"{best['Analysis']} - {best['Model']}",
                "Accuracy": best["Accuracy"],
                "F1": best["F1"],
                "AUC": best["AUC"],
                "MCC": best["MCC"],
                "Sample Rate": best.get("Sample_Rate", "N/A"),
            })

    # 2. Best combined model per feature set
    combined_df = df[df["Analysis"].str.contains("Combined", na=False)]
    if not combined_df.empty:
        for fs in combined_df["Feature_Set"].unique():
            subset = combined_df[combined_df["Feature_Set"] == fs]
            best_idx = subset["AUC"].idxmax()
            best = subset.loc[best_idx]
            rows.append({
                "Category": "Best Combined Model",
                "Feature Set": fs,
                "Detail": f"{best['Model']}",
                "Accuracy": best["Accuracy"],
                "F1": best["F1"],
                "AUC": best["AUC"],
                "MCC": best["MCC"],
                "Sample Rate": best.get("Sample_Rate", "N/A"),
            })

    # 3. Cross-dataset generalization (best)
    cross_df = df[~df["Analysis"].str.contains("Self|Combined", na=False)]
    if not cross_df.empty:
        for fs in cross_df["Feature_Set"].unique():
            subset = cross_df[cross_df["Feature_Set"] == fs]
            best_idx = subset["AUC"].idxmax()
            best = subset.loc[best_idx]
            rows.append({
                "Category": "Best Cross-Dataset",
                "Feature Set": fs,
                "Detail": f"{best['Analysis']} - {best['Model']}",
                "Accuracy": best["Accuracy"],
                "F1": best["F1"],
                "AUC": best["AUC"],
                "MCC": best["MCC"],
                "Sample Rate": best.get("Sample_Rate", "N/A"),
            })

    # 4. Model comparison across all experiments
    model_stats = df.groupby("Model").agg(
        Mean_Accuracy=("Accuracy", "mean"),
        Mean_F1=("F1", "mean"),
        Mean_AUC=("AUC", "mean"),
        Mean_MCC=("MCC", "mean"),
        Experiments=("Experiment", "nunique"),
    ).reset_index()
    for _, r in model_stats.iterrows():
        rows.append({
            "Category": "Model Overall Average",
            "Feature Set": "All",
            "Detail": r["Model"],
            "Accuracy": r["Mean_Accuracy"],
            "F1": r["Mean_F1"],
            "AUC": r["Mean_AUC"],
            "MCC": r["Mean_MCC"],
            "Sample Rate": f"{int(r['Experiments'])} experiments",
        })

    summary_df = pd.DataFrame(rows)
    summary_df.to_excel(writer, sheet_name="Analytical Summary", index=False)


def format_workbook(path):
    """Apply formatting to all sheets."""
    wb = load_workbook(path)
    for ws in wb.worksheets:
        max_col = ws.max_column
        max_row = ws.max_row
        if max_row < 1 or max_col < 1:
            continue

        # Style header
        style_header_row(ws, 1, max_col)

        # Format numeric cells
        for row in range(2, max_row + 1):
            for col in range(1, max_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal="center")
                header = ws.cell(row=1, column=col).value
                if header and isinstance(cell.value, (int, float)):
                    h = str(header).lower()
                    if h in ("accuracy", "f1", "auc", "mcc", "mean_accuracy",
                             "mean_f1", "mean_auc", "mean_mcc",
                             "auc_ci_lo", "auc_ci_hi", "p_value"):
                        cell.number_format = NUM_FMT_DEC

        # Highlight best/worst AUC per group in Raw Data & Key Metrics
        if ws.title in ("Raw Data", "Key Metrics", "Analytical Summary"):
            auc_col = None
            for c in range(1, max_col + 1):
                if ws.cell(row=1, column=c).value and "AUC" == str(ws.cell(row=1, column=c).value).upper().strip():
                    auc_col = c
                    break
            if auc_col and max_row > 1:
                vals = []
                for r in range(2, max_row + 1):
                    v = ws.cell(row=r, column=auc_col).value
                    if isinstance(v, (int, float)):
                        vals.append((r, v))
                if vals:
                    best_row = max(vals, key=lambda x: x[1])[0]
                    worst_row = min(vals, key=lambda x: x[1])[0]
                    for c in range(1, max_col + 1):
                        ws.cell(row=best_row, column=c).fill = BEST_FILL
                        ws.cell(row=worst_row, column=c).fill = WORST_FILL

        auto_width(ws)

    # Freeze panes on all sheets
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"

    wb.save(path)


def main():
    print("Loading results...")
    df = load_all_csv_results()
    rankings = load_ranking_data()

    if df.empty:
        print("ERROR: No CSV results found in", RESULTS_DIR)
        return

    print(f"Found {len(df)} result rows across {df['Experiment'].nunique()} experiments")
    print(f"Found {len(rankings)} ranking files")

    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        print("Writing Raw Data...")
        write_raw_data(writer, df)

        print("Writing Key Metrics...")
        write_key_metrics(writer, df)

        print("Writing Cross-Dataset Analysis...")
        write_cross_dataset(writer, df)

        print("Writing Analytical Summary...")
        write_analytical_summary(writer, df)

        print("Writing Feature Rankings...")
        write_feature_rankings(writer, rankings)

    print("Formatting workbook...")
    format_workbook(OUTPUT_FILE)

    print(f"\n✅ Report saved to: {OUTPUT_FILE}")
    print(f"   Sheets: Raw Data, Key Metrics, Cross-Dataset, Analytical Summary, + {len(rankings)} ranking sheets")


if __name__ == "__main__":
    main()
