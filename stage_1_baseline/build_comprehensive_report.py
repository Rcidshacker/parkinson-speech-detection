import os, json, glob
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

RESULTS_DIR = os.path.join(os.path.dirname(__file__), "results")
OUTPUT_FILE = os.path.join(os.path.dirname(__file__), "comprehensive_analysis_report.xlsx")

# Formatting defined by user
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
BEST_FILL = PatternFill("solid", fgColor="C6EFCE")
BEST_FONT = Font(color="006100")
WORST_FILL = PatternFill("solid", fgColor="FFC7CE")
WORST_FONT = Font(color="9C0006")
NEUTRAL_FILL_1 = PatternFill("solid", fgColor="FFFFFF")
NEUTRAL_FILL_2 = PatternFill("solid", fgColor="F2F2F2")
DIVIDER_FILL = PatternFill("solid", fgColor="FFEB9C")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

def format_sheet(ws, is_raw=False):
    # Auto-fit columns
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = min(adjusted_width, 60)
    
    # Header format
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER
        
    # Freeze row 1
    ws.freeze_panes = "A2"
    
    # Alternate row colors and borders for data
    if ws.max_row > 1:
        for r_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column)):
            fill = NEUTRAL_FILL_1 if r_idx % 2 == 0 else NEUTRAL_FILL_2
            for cell in row:
                if not cell.fill or cell.fill.fgColor.rgb == "00000000":
                    cell.fill = fill
                cell.border = THIN_BORDER
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if isinstance(cell.value, float):
                    cell.number_format = "0.0000"

def highlight_best_worst(ws, metric_cols):
    """Highlight max (green) and min (red) in specified numeric columns."""
    if ws.max_row < 2: return
    for col_idx in metric_cols:
        col_letter = get_column_letter(col_idx)
        vals = []
        for r in range(2, ws.max_row + 1):
            val = ws[f"{col_letter}{r}"].value
            if isinstance(val, (int, float)):
                vals.append((r, val))
        if vals:
            max_r = max(vals, key=lambda x: x[1])[0]
            min_r = min(vals, key=lambda x: x[1])[0]
            
            ws[f"{col_letter}{max_r}"].fill = BEST_FILL
            ws[f"{col_letter}{max_r}"].font = BEST_FONT
            ws[f"{col_letter}{min_r}"].fill = WORST_FILL
            ws[f"{col_letter}{min_r}"].font = WORST_FONT

def parse_directories():
    all_data = []
    visuals = []
    
    dirs = [d for d in glob.glob(os.path.join(RESULTS_DIR, "*")) if os.path.isdir(d) and os.path.basename(d) not in ["final", "archive"]]
    
    for d in dirs:
        exp_name = os.path.basename(d)
        
        # Parse CSV
        csv_file = os.path.join(d, "csv_results", "analysis_summary.csv")
        csv_df = None
        if os.path.exists(csv_file):
            csv_df = pd.read_csv(csv_file)
            for _, row in csv_df.iterrows():
                row_dict = row.to_dict()
                row_dict["Source_Experiment"] = exp_name
                all_data.append(row_dict)
                
        # Parse JSON
        json_file = os.path.join(d, "reports", "detailed_results.json")
        json_data = {}
        if os.path.exists(json_file):
            try:
                with open(json_file, 'r', encoding='utf-8') as f:
                    json_data = json.load(f)
            except:
                pass

        # Parse visual info (images)
        for img in glob.glob(os.path.join(d, "visualizations", "*.png")) + glob.glob(os.path.join(d, "matrices", "*.png")):
            img_name = os.path.basename(img)
            desc = ""
            
            # Extract basic metric to add context to image
            context_metric = ""
            
            # Simple heuristic mapping
            if "_cm." in img_name:
                desc = f"Confusion matrix for evaluating the predictive performance. Shows the prediction distribution across classes in {exp_name}."
            elif "_ROC" in img_name:
                desc = f"Receiver Operating Characteristic (ROC) curve showing diagnostic capability. Higher Area Under Curve (AUC) indicates superior performance for {exp_name}."
            elif "heatmap" in img_name.lower():
                desc = f"Performance heatmap illustrating metric correlations or values across conditions in {exp_name}."
            else:
                desc = f"Visual chart representation for the experiment {exp_name}."
                
            visuals.append({
                "Source_Experiment": exp_name,
                "Image_Name": img_name,
                "Type": "Confusion Matrix" if "_cm." in img_name else ("ROC Curve" if "roc" in img_name.lower() else "Heatmap"),
                "Auto_Generated_Summary": desc
            })

    raw_df = pd.DataFrame(all_data) if all_data else pd.DataFrame()
    vis_df = pd.DataFrame(visuals) if visuals else pd.DataFrame()
    
    if not raw_df.empty:
        raw_df.fillna("", inplace=True)
    if not vis_df.empty:
        vis_df.fillna("", inplace=True)
        
    return raw_df, vis_df

def build_report():
    print("Parsing experiments...")
    raw_df, vis_df = parse_directories()
    
    if raw_df.empty:
        print("No data found!")
        return

    wb = Workbook()
    
    # 1. Overview_Summary
    ws_over = wb.active
    ws_over.title = "Overview_Summary"
    overview = raw_df.groupby("Source_Experiment").agg(
        Total_Models=("Model", "count"),
        Avg_Accuracy=("Accuracy", "mean"),
        Max_AUC=("AUC", "max")
    ).reset_index()
    for r in dataframe_to_rows(overview, index=False, header=True):
        ws_over.append(r)
        
    # 2. Detailed_Metrics
    ws_det = wb.create_sheet("Detailed_Metrics")
    detailed = raw_df[["Source_Experiment", "Analysis", "Model", "Accuracy", "F1", "AUC", "MCC"]].copy()
    for r in dataframe_to_rows(detailed, index=False, header=True):
        ws_det.append(r)
        
    # 3. Statistical_Analysis
    ws_stat = wb.create_sheet("Statistical_Analysis")
    numeric_cols = ["Accuracy", "F1", "AUC", "MCC"]
    # Ensure columns are numeric
    for c in numeric_cols:
        if c in raw_df.columns:
            raw_df[c] = pd.to_numeric(raw_df[c], errors='coerce')
            
    stats = raw_df.groupby("Source_Experiment")[numeric_cols].agg(["mean", "std", "min", "max"]).reset_index()
    stats.columns = ['_'.join(col).strip() if col[1] else col[0] for col in stats.columns.values]
    for r in dataframe_to_rows(stats, index=False, header=True):
        ws_stat.append(r)
        
    # 4. Comparison_Matrix
    ws_comp = wb.create_sheet("Comparison_Matrix")
    try:
        comp = raw_df.pivot_table(index="Model", columns="Analysis", values="AUC", aggfunc="mean").reset_index()
        comp.fillna("", inplace=True)
        for r in dataframe_to_rows(comp, index=False, header=True):
            ws_comp.append(r)
    except Exception as e:
        ws_comp.append(["Error pivoting Comparison Matrix", str(e)])
        
    # 5. Rankings
    ws_rank = wb.create_sheet("Rankings")
    ranked = raw_df.sort_values(by="AUC", ascending=False).reset_index(drop=True)
    ranked["Global_Rank"] = ranked.index + 1
    ranked = ranked[["Global_Rank", "Source_Experiment", "Analysis", "Model", "AUC", "Accuracy", "F1", "MCC"]]
    for r in dataframe_to_rows(ranked, index=False, header=True):
        ws_rank.append(r)
        
    # 6. Raw_Data
    ws_raw = wb.create_sheet("Raw_Data")
    for r in dataframe_to_rows(raw_df, index=False, header=True):
        ws_raw.append(r)
        
    # 7. Visual_Summary
    ws_vis = wb.create_sheet("Visual_Summary")
    if not vis_df.empty:
        for r in dataframe_to_rows(vis_df, index=False, header=True):
            ws_vis.append(r)
    else:
        ws_vis.append(["Source_Experiment", "Image_Name", "Type", "Auto_Generated_Summary"])
        ws_vis.append(["None", "None", "None", "No images found."])

    # Formatting all sheets
    sheet_metric_cols = {
        "Overview_Summary": [2, 3, 4], 
        "Detailed_Metrics": [4, 5, 6, 7],
        "Statistical_Analysis": [], 
        "Comparison_Matrix": list(range(2, len(comp.columns)+1)) if 'comp' in locals() else [],
        "Rankings": [5, 6, 7, 8],
        "Raw_Data": [raw_df.columns.get_loc(c)+1 for c in ["Accuracy", "F1", "AUC", "MCC"] if c in raw_df.columns],
        "Visual_Summary": []
    }
    
    for title, metric_cols in sheet_metric_cols.items():
        if title in wb.sheetnames:
            ws = wb[title]
            format_sheet(ws)
            if metric_cols:
                highlight_best_worst(ws, metric_cols)
                
    # Add dividers in Raw_Data between different experiments
    ws_raw = wb["Raw_Data"]
    exp_col = 1
    for c_idx in range(1, ws_raw.max_column + 1):
        if ws_raw.cell(row=1, column=c_idx).value == "Source_Experiment":
            exp_col = c_idx
            break
            
    last_exp = None
    for r in range(2, ws_raw.max_row + 1):
        curr_exp = ws_raw.cell(row=r, column=exp_col).value
        # If experiment changes, insert a color band logic or a strong border
        if curr_exp and curr_exp != last_exp and last_exp is not None:
             for c in range(1, ws_raw.max_column + 1):
                 # Add top divider border
                 current_border = ws_raw.cell(row=r, column=c).border
                 ws_raw.cell(row=r, column=c).border = Border(
                     top=Side(style="thick", color="D6E4F0"),
                     bottom=current_border.bottom,
                     left=current_border.left,
                     right=current_border.right
                 )
        last_exp = curr_exp

    print(f"Saving report to {OUTPUT_FILE}...")
    wb.save(OUTPUT_FILE)
    print("Done!")

if __name__ == "__main__":
    build_report()
