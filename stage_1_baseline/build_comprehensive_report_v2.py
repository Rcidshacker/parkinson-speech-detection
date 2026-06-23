"""
build_comprehensive_report_v2.py
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Forensically corrected rebuild of build_comprehensive_report.py.

FIXES APPLIED
─────────────
FIX 1  Silent failure → Explicit loading log with missing-file manifest
FIX 2  Overview_Summary now STRATIFIED by domain type (Self / Cross-Domain)
        instead of a single pooled Avg_Accuracy + cherry-picked Max_AUC
FIX 3  highlight_best_worst handles MCC correctly:
          MCC ≥ 0  →  higher = green  (same as other metrics)
          MCC < 0  →  always red (worse than random, regardless of rank)
FIX 4  New "Data_Integrity" sheet: flags duplicate experiments, warns user
FIX 5  AUC Confidence Intervals surfaced in Overview (not just buried in Raw)
FIX 6  Visual_Summary now distinguishes image type by filename more robustly
        and marks any unclassified images explicitly instead of defaulting

Author : Forensic Data Architect (remediation pass)
Date   : 2026-03-28
"""

import os
import json
import glob
import logging
import hashlib
from collections import defaultdict

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

# ─────────────────────────── PATHS ──────────────────────────────────────────
RESULTS_DIR = os.path.join(os.path.dirname(__file__), "results")
OUTPUT_FILE = os.path.join(
    os.path.dirname(__file__), "comprehensive_analysis_report_v2.xlsx"
)
EXCLUDED_DIRS = {"final", "archive"}

# ─────────────────────────── LOGGING ────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("report_builder")

# ─────────────────────────── DOMAIN CLASSIFICATION ───────────────────────────
SELF_DOMAIN_LABELS    = {"pc_gita (Self)", "voice_dataset (Self)"}
CROSS_DOMAIN_LABELS   = {"pc_gita→voice_dataset", "voice_dataset→pc_gita"}
COMBINED_LABEL        = "Combined (All)"

def domain_type(analysis_label: str) -> str:
    if analysis_label in SELF_DOMAIN_LABELS:
        return "Self"
    if analysis_label in CROSS_DOMAIN_LABELS:
        return "Cross-Domain"
    return "Combined"

# ─────────────────────────── STYLE CONSTANTS ─────────────────────────────────
HEADER_FILL    = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT    = Font(bold=True, color="FFFFFF", size=11)

BEST_FILL      = PatternFill("solid", fgColor="C6EFCE")
BEST_FONT      = Font(color="006100", bold=True)

WORST_FILL     = PatternFill("solid", fgColor="FFC7CE")
WORST_FONT     = Font(color="9C0006", bold=True)

WARN_FILL      = PatternFill("solid", fgColor="FFEB9C")   # amber for negative MCC
WARN_FONT      = Font(color="9C5700", bold=True)

DUPE_FILL      = PatternFill("solid", fgColor="FCE4D6")   # salmon for duplicates
DUPE_FONT      = Font(color="833C00")

SELF_FILL      = PatternFill("solid", fgColor="DDEEFF")   # blue tint for Self rows
CROSS_FILL     = PatternFill("solid", fgColor="FFF2CC")   # yellow tint for Cross rows
COMBINED_FILL  = PatternFill("solid", fgColor="E2EFDA")   # green tint for Combined

NEUTRAL_1      = PatternFill("solid", fgColor="FFFFFF")
NEUTRAL_2      = PatternFill("solid", fgColor="F5F5F5")

THIN           = Side(style="thin")
THICK_TOP      = Side(style="medium", color="4472C4")
THIN_BORDER    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ─────────────────────────── HELPERS ─────────────────────────────────────────
def col_letter(idx: int) -> str:
    return get_column_letter(idx)


def format_sheet_base(ws):
    """Apply header styling, freeze pane, alternating rows, borders."""
    # Header row
    for cell in ws[1]:
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        cell.border    = THIN_BORDER

    ws.freeze_panes = "A2"

    # Data rows – alternate neutral fills, centre-align, number format
    for r_idx, row in enumerate(
        ws.iter_rows(min_row=2, max_row=ws.max_row,
                     min_col=1, max_col=ws.max_column)
    ):
        fill = NEUTRAL_1 if r_idx % 2 == 0 else NEUTRAL_2
        for cell in row:
            if not cell.fill or cell.fill.patternType is None:
                cell.fill = fill
            cell.border    = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if isinstance(cell.value, float):
                cell.number_format = "0.0000"

    # Auto-fit columns (cap at 60)
    for col in ws.columns:
        max_len = 0
        col_ltr = col[0].column_letter
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_len:
                    max_len = len(str(cell.value))
            except Exception:
                pass
        ws.column_dimensions[col_ltr].width = min(max_len + 4, 60)


def highlight_best_worst(ws, metric_cols: list, mcc_cols: list = None):
    """
    FIX 3 — Correct highlight logic:
      • For all normal metrics (Accuracy, F1, AUC):
            max → green,  min → red
      • For MCC specifically:
            max (if ≥ 0) → green
            any value < 0 → always red (worse than random)
            min among non-negative → no special treatment unless it is also
            the global min
    """
    if mcc_cols is None:
        mcc_cols = []

    for col_idx in metric_cols:
        cl = col_letter(col_idx)
        vals = []
        for r in range(2, ws.max_row + 1):
            v = ws[f"{cl}{r}"].value
            if isinstance(v, (int, float)):
                vals.append((r, v))
        if not vals:
            continue

        is_mcc = col_idx in mcc_cols

        if is_mcc:
            # Paint every negative MCC cell red unconditionally
            for (r, v) in vals:
                if v < 0:
                    ws[f"{cl}{r}"].fill = WORST_FILL
                    ws[f"{cl}{r}"].font = WARN_FONT

            # Best non-negative value gets green
            pos_vals = [(r, v) for (r, v) in vals if v >= 0]
            if pos_vals:
                best_r = max(pos_vals, key=lambda x: x[1])[0]
                ws[f"{cl}{best_r}"].fill = BEST_FILL
                ws[f"{cl}{best_r}"].font = BEST_FONT

        else:
            # Standard: max → green, min → red
            best_r  = max(vals, key=lambda x: x[1])[0]
            worst_r = min(vals, key=lambda x: x[1])[0]
            ws[f"{cl}{best_r}"].fill  = BEST_FILL
            ws[f"{cl}{best_r}"].font  = BEST_FONT
            ws[f"{cl}{worst_r}"].fill = WORST_FILL
            ws[f"{cl}{worst_r}"].font = WORST_FONT


def colour_domain_rows(ws, analysis_col_idx: int):
    """Colour-band rows by Self / Cross-Domain / Combined."""
    for r in range(2, ws.max_row + 1):
        label = ws.cell(row=r, column=analysis_col_idx).value or ""
        dtype = domain_type(label)
        fill = (SELF_FILL if dtype == "Self"
                else CROSS_FILL if dtype == "Cross-Domain"
                else COMBINED_FILL)
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).fill = fill


# ─────────────────────────── DATA LOADING ────────────────────────────────────
def csv_fingerprint(path: str) -> str:
    """SHA-256 of the raw CSV bytes — used to detect true duplicates."""
    with open(path, "rb") as f:
        return hashlib.sha256(f.read()).hexdigest()


def parse_directories():
    """
    FIX 1 — Explicit loading log: every missing file is recorded.
    Returns: (raw_df, vis_df, load_log_df)
    """
    all_data   = []
    visuals    = []
    load_log   = []          # one row per experiment, tracking what was found

    dirs = sorted([
        d for d in glob.glob(os.path.join(RESULTS_DIR, "*"))
        if os.path.isdir(d)
        and os.path.basename(d) not in EXCLUDED_DIRS
    ])

    log.info(f"Found {len(dirs)} experiment directories.")

    for d in dirs:
        exp_name  = os.path.basename(d)
        csv_file  = os.path.join(d, "csv_results", "analysis_summary.csv")
        json_file = os.path.join(d, "reports",     "detailed_results.json")

        entry = {
            "Experiment"      : exp_name,
            "CSV_Found"       : os.path.exists(csv_file),
            "JSON_Found"      : os.path.exists(json_file),
            "CSV_Rows_Loaded" : 0,
            "CSV_Path"        : csv_file,
            "Error"           : "",
        }

        # ── CSV ──────────────────────────────────────────────────────────────
        if entry["CSV_Found"]:
            try:
                csv_df = pd.read_csv(csv_file)
                entry["CSV_Rows_Loaded"] = len(csv_df)
                for _, row in csv_df.iterrows():
                    rd = row.to_dict()
                    rd["Source_Experiment"] = exp_name
                    rd["Domain_Type"]       = domain_type(rd.get("Analysis", ""))
                    all_data.append(rd)
                log.info(f"  ✓  {exp_name}: loaded {len(csv_df)} rows")
            except Exception as e:
                entry["Error"] = f"CSV parse error: {e}"
                log.error(f"  ✗  {exp_name}: CSV parse failed — {e}")
        else:
            entry["Error"] = "csv_results/analysis_summary.csv NOT FOUND"
            log.warning(f"  ✗  {exp_name}: missing CSV")

        # ── JSON (informational – used for integrity checks) ──────────────
        if entry["JSON_Found"]:
            try:
                with open(json_file, "r", encoding="utf-8") as f:
                    json.load(f)          # validate JSON is well-formed
            except json.JSONDecodeError as e:
                entry["Error"] += f" | JSON decode error: {e}"
                log.error(f"  ✗  {exp_name}: JSON malformed — {e}")
            except Exception as e:
                entry["Error"] += f" | JSON read error: {e}"
                log.error(f"  ✗  {exp_name}: JSON unreadable — {e}")
        else:
            log.warning(f"  ✗  {exp_name}: missing JSON")

        load_log.append(entry)

        # ── Images ───────────────────────────────────────────────────────────
        imgs = (
            glob.glob(os.path.join(d, "visualizations", "*.png")) +
            glob.glob(os.path.join(d, "matrices",       "*.png"))
        )
        for img in imgs:
            img_name = os.path.basename(img)
            name_lower = img_name.lower()

            if "_cm." in img_name:
                img_type = "Confusion Matrix"
                summary  = (f"Confusion matrix showing prediction distribution "
                            f"for {exp_name}.")
            elif "roc" in name_lower:
                img_type = "ROC Curve"
                summary  = (f"ROC curve showing classifier discrimination for "
                            f"{exp_name}.")
            elif "heatmap" in name_lower:
                img_type = "Heatmap"
                summary  = (f"Performance heatmap across models/conditions for "
                            f"{exp_name}.")
            else:
                # FIX 6 — explicit UNKNOWN instead of defaulting silently
                img_type = "UNKNOWN — review manually"
                summary  = (f"Image {img_name} could not be auto-classified. "
                            f"Manual inspection required.")
                log.warning(f"  ⚠  Unclassified image: {img_name}")

            visuals.append({
                "Source_Experiment"    : exp_name,
                "Image_Name"           : img_name,
                "Type"                 : img_type,
                "Auto_Generated_Summary": summary,
            })

    raw_df  = pd.DataFrame(all_data)  if all_data  else pd.DataFrame()
    vis_df  = pd.DataFrame(visuals)   if visuals   else pd.DataFrame()
    log_df  = pd.DataFrame(load_log)

    if not raw_df.empty:
        for c in ["Accuracy", "F1", "AUC", "AUC_CI_lo", "AUC_CI_hi", "MCC"]:
            if c in raw_df.columns:
                raw_df[c] = pd.to_numeric(raw_df[c], errors="coerce")

    return raw_df, vis_df, log_df


# ─────────────────────────── DUPLICATE DETECTION ─────────────────────────────
def detect_duplicates(raw_df: pd.DataFrame, load_log_df: pd.DataFrame) -> pd.DataFrame:
    """
    FIX 4 — Duplicate detection.
    Two experiments are flagged as duplicates when their per-row metric vectors
    (for matching Analysis × Model pairs) are identical within 1e-6 tolerance.
    Falls back to CSV hash comparison where available.
    """
    records = []
    exps = sorted(raw_df["Source_Experiment"].unique())

    # Group by experiment → dict of {(Analysis,Model): metrics_tuple}
    exp_vectors = {}
    for exp in exps:
        subset = raw_df[raw_df["Source_Experiment"] == exp].copy()
        vec = {}
        for _, row in subset.iterrows():
            key = (row.get("Analysis",""), row.get("Model",""))
            vec[key] = (
                round(float(row.get("Accuracy", 0)), 8),
                round(float(row.get("AUC",      0)), 8),
                round(float(row.get("MCC",      0)), 8),
            )
        exp_vectors[exp] = vec

    seen_pairs = set()
    for i, e1 in enumerate(exps):
        for e2 in exps[i+1:]:
            pair = tuple(sorted([e1, e2]))
            if pair in seen_pairs:
                continue
            seen_pairs.add(pair)

            v1, v2 = exp_vectors[e1], exp_vectors[e2]
            common_keys = set(v1) & set(v2)
            if not common_keys:
                continue

            diffs = sum(
                any(abs(a-b) > 1e-6 for a,b in zip(v1[k], v2[k]))
                for k in common_keys
            )
            match_pct = 100.0 * (1 - diffs / len(common_keys))

            if match_pct >= 95.0:
                # Try hash-level confirmation
                csv1 = load_log_df.loc[
                    load_log_df["Experiment"] == e1, "CSV_Path"
                ].values
                csv2 = load_log_df.loc[
                    load_log_df["Experiment"] == e2, "CSV_Path"
                ].values
                hash_match = "N/A"
                if csv1.size and csv2.size:
                    try:
                        h1 = csv_fingerprint(csv1[0])
                        h2 = csv_fingerprint(csv2[0])
                        hash_match = "IDENTICAL" if h1 == h2 else "Near-identical (re-run)"
                    except Exception:
                        hash_match = "Hash check failed"

                records.append({
                    "Experiment_A"      : e1,
                    "Experiment_B"      : e2,
                    "Metric_Match_Pct"  : round(match_pct, 2),
                    "File_Hash_Compare" : hash_match,
                    "Keys_Compared"     : len(common_keys),
                    "Differences"       : diffs,
                    "Verdict"           : (
                        "⚠ LIKELY DUPLICATE — same config re-run. "
                        "Consider averaging or dropping one."
                    ),
                })
                log.warning(
                    f"  ⚠  Duplicate detected: {e1[-25:]} ↔ {e2[-25:]} "
                    f"({match_pct:.1f}% match, {hash_match})"
                )

    return pd.DataFrame(records) if records else pd.DataFrame(
        columns=["Experiment_A","Experiment_B","Metric_Match_Pct",
                 "File_Hash_Compare","Keys_Compared","Differences","Verdict"]
    )


# ─────────────────────────── OVERVIEW BUILDER ────────────────────────────────
def build_stratified_overview(raw_df: pd.DataFrame) -> pd.DataFrame:
    """
    FIX 2 — Stratified Overview.
    Instead of one pooled Avg_Accuracy + cherry-picked Max_AUC per experiment,
    we now produce one row per (Experiment × Domain_Type) combination, with:
      - N_Models               : how many models ran this condition
      - Mean_AUC (±std)        : honest central tendency
      - Median_AUC             : robust to outliers
      - Best_AUC               : max within domain type only
      - Mean_AUC_CI_lo/hi      : mean of the per-model CI bounds (proxy)
      - Mean_Accuracy
      - Mean_MCC
      - Any_Negative_MCC       : flag for investigators
    """
    rows = []
    for (exp, dtype), grp in raw_df.groupby(
        ["Source_Experiment", "Domain_Type"]
    ):
        neg_mcc = (grp["MCC"] < 0).any() if "MCC" in grp.columns else False
        rows.append({
            "Source_Experiment"  : exp,
            "Domain_Type"        : dtype,
            "N_Models"           : len(grp),
            "Mean_AUC"           : grp["AUC"].mean(),
            "Std_AUC"            : grp["AUC"].std(),
            "Median_AUC"         : grp["AUC"].median(),
            "Best_AUC"           : grp["AUC"].max(),
            "Worst_AUC"          : grp["AUC"].min(),
            "Mean_AUC_CI_lo"     : grp["AUC_CI_lo"].mean() if "AUC_CI_lo" in grp.columns else None,
            "Mean_AUC_CI_hi"     : grp["AUC_CI_hi"].mean() if "AUC_CI_hi" in grp.columns else None,
            "Mean_Accuracy"      : grp["Accuracy"].mean(),
            "Mean_MCC"           : grp["MCC"].mean()        if "MCC" in grp.columns else None,
            "Any_Negative_MCC"   : "YES ⚠" if neg_mcc else "no",
        })

    df = pd.DataFrame(rows)
    # Sort: experiment name, then Self first, Cross-Domain second, Combined last
    order = {"Self": 0, "Cross-Domain": 1, "Combined": 2}
    df["_sort"] = df["Domain_Type"].map(order).fillna(3)
    df = df.sort_values(["Source_Experiment", "_sort"]).drop(columns=["_sort"])
    return df.reset_index(drop=True)


# ─────────────────────────── CI SUMMARY SHEET ────────────────────────────────
def build_ci_summary(raw_df: pd.DataFrame) -> pd.DataFrame:
    """
    FIX 5 — Explicit AUC Confidence Interval sheet.
    One row per (Experiment × Analysis × Model) with:
      AUC, CI_lo, CI_hi, CI_width, CI_width_flag
    A wide CI flags small-sample uncertainty.
    """
    cols = ["Source_Experiment", "Analysis", "Model",
            "AUC", "AUC_CI_lo", "AUC_CI_hi"]
    available = [c for c in cols if c in raw_df.columns]
    df = raw_df[available].copy()

    if "AUC_CI_lo" in df.columns and "AUC_CI_hi" in df.columns:
        df["CI_Width"] = df["AUC_CI_hi"] - df["AUC_CI_lo"]
        # Flag: CI width > 0.15 = unreliable (small test set)
        df["Reliability"] = df["CI_Width"].apply(
            lambda w: "⚠ WIDE — small test set?" if (
                isinstance(w, float) and w > 0.15
            ) else ("OK" if isinstance(w, float) else "N/A")
        )

    df["Domain_Type"] = df["Analysis"].apply(domain_type) if "Analysis" in df.columns else ""
    df = df.sort_values(
        ["Source_Experiment", "Domain_Type", "Analysis", "AUC"],
        ascending=[True, True, True, False]
    ).reset_index(drop=True)
    return df


# ─────────────────────────── CROSS-DOMAIN COLLAPSE SHEET ─────────────────────
def build_generalisation_verdict(raw_df: pd.DataFrame) -> pd.DataFrame:
    """
    Extra sheet: per-experiment, per-model generalisation gap.
    Gap = Self AUC − Cross-Domain AUC. Large gap = dataset-specific learning.
    """
    rows = []
    for exp, grp in raw_df.groupby("Source_Experiment"):
        self_rows  = grp[grp["Domain_Type"] == "Self"]
        cross_rows = grp[grp["Domain_Type"] == "Cross-Domain"]

        for model in grp["Model"].unique():
            self_auc  = self_rows[self_rows["Model"]==model]["AUC"].mean()
            cross_auc = cross_rows[cross_rows["Model"]==model]["AUC"].mean()

            if pd.notna(self_auc) and pd.notna(cross_auc):
                gap = self_auc - cross_auc
                rows.append({
                    "Source_Experiment"  : exp,
                    "Model"              : model,
                    "Mean_Self_AUC"      : round(self_auc,  4),
                    "Mean_Cross_AUC"     : round(cross_auc, 4),
                    "Generalisation_Gap" : round(gap,       4),
                    "Verdict"            : (
                        "✗ SEVERE — model learns dataset artifact"  if gap > 0.25 else
                        "⚠ MODERATE — partial transfer"             if gap > 0.12 else
                        "✓ ACCEPTABLE"
                    ),
                })

    df = pd.DataFrame(rows)
    if not df.empty:
        df = df.sort_values("Generalisation_Gap", ascending=False).reset_index(drop=True)
    return df


# ─────────────────────────── MAIN BUILD ──────────────────────────────────────
def build_report():
    log.info("=" * 68)
    log.info("build_comprehensive_report_v2.py  —  Forensic Edition")
    log.info("=" * 68)

    log.info("\nPHASE 1 — Loading experiments...")
    raw_df, vis_df, load_log_df = parse_directories()

    if raw_df.empty:
        log.error("No data loaded — aborting.")
        return

    # ── Summary statistics ────────────────────────────────────────────────────
    missing_csv = load_log_df[~load_log_df["CSV_Found"]]
    if not missing_csv.empty:
        log.warning(
            f"\n{'─'*60}\n"
            f"  {len(missing_csv)} experiments have MISSING CSV files:\n" +
            "\n".join(f"  • {r}" for r in missing_csv["Experiment"]) +
            f"\n{'─'*60}"
        )

    log.info(f"\nTotal rows loaded : {len(raw_df)}")
    log.info(f"Unique experiments: {raw_df['Source_Experiment'].nunique()}")
    log.info(f"Models found      : {sorted(raw_df['Model'].unique())}")
    log.info(f"Analysis types    : {sorted(raw_df['Analysis'].unique())}")

    log.info("\nPHASE 2 — Detecting duplicates...")
    dupe_df = detect_duplicates(raw_df, load_log_df)
    log.info(f"  → {len(dupe_df)} duplicate pair(s) detected.")

    log.info("\nPHASE 3 — Building Excel workbook...")
    wb = Workbook()

    # ═══════════════════════════════════════════════════════════════════════
    # SHEET 1  —  Data_Integrity  (FIX 4) — FIRST so it is seen immediately
    # ═══════════════════════════════════════════════════════════════════════
    ws_integrity = wb.active
    ws_integrity.title = "Data_Integrity"

    # Loading log sub-table
    ws_integrity.append(["LOADING MANIFEST"])
    ws_integrity["A1"].font = Font(bold=True, size=13, color="1F4E79")
    ws_integrity.append([])

    load_cols = ["Experiment","CSV_Found","JSON_Found","CSV_Rows_Loaded","Error"]
    ws_integrity.append(load_cols)
    for _, row in load_log_df[load_cols].iterrows():
        ws_integrity.append(list(row))

    ws_integrity.append([])
    ws_integrity.append([])

    # Duplicate manifest sub-table
    start_row = ws_integrity.max_row + 1
    ws_integrity.cell(row=start_row, column=1,
                      value="DUPLICATE EXPERIMENT MANIFEST").font = Font(
                          bold=True, size=13, color="9C0006")
    ws_integrity.append([])

    dupe_cols = ["Experiment_A","Experiment_B","Metric_Match_Pct",
                 "File_Hash_Compare","Keys_Compared","Differences","Verdict"]
    ws_integrity.append(dupe_cols)
    if dupe_df.empty:
        ws_integrity.append(["No duplicates detected."] + [""] * 6)
    else:
        for _, row in dupe_df[dupe_cols].iterrows():
            ws_integrity.append(list(row))
            # Paint the row salmon
            r = ws_integrity.max_row
            for c in range(1, len(dupe_cols)+1):
                ws_integrity.cell(row=r, column=c).fill = DUPE_FILL
                ws_integrity.cell(row=r, column=c).font = DUPE_FONT

    format_sheet_base(ws_integrity)

    # ═══════════════════════════════════════════════════════════════════════
    # SHEET 2  —  Overview_Summary  (FIX 2 — Stratified)
    # ═══════════════════════════════════════════════════════════════════════
    ws_over = wb.create_sheet("Overview_Summary")
    overview_df = build_stratified_overview(raw_df)
    for r in dataframe_to_rows(overview_df, index=False, header=True):
        ws_over.append(r)

    format_sheet_base(ws_over)

    # Colour-band by domain type (col 2 = Domain_Type)
    for r in range(2, ws_over.max_row + 1):
        dtype = ws_over.cell(row=r, column=2).value or ""
        fill  = (SELF_FILL   if dtype == "Self"         else
                 CROSS_FILL  if dtype == "Cross-Domain" else
                 COMBINED_FILL)
        for c in range(1, ws_over.max_column + 1):
            ws_over.cell(row=r, column=c).fill = fill

    # Highlight best/worst within each domain group
    col_map = {name: i+1 for i, name in enumerate(overview_df.columns)}
    metric_cols_over = [col_map[c] for c in
                        ["Mean_AUC","Median_AUC","Best_AUC","Mean_Accuracy","Mean_MCC"]
                        if c in col_map]
    mcc_cols_over    = [col_map["Mean_MCC"]] if "Mean_MCC" in col_map else []
    highlight_best_worst(ws_over, metric_cols_over, mcc_cols_over)

    # Add a legend note
    ws_over.cell(row=ws_over.max_row + 2, column=1,
                 value="Legend: Blue = Self-domain | Yellow = Cross-Domain | Green = Combined").font = Font(
                     italic=True, color="595959")

    # ═══════════════════════════════════════════════════════════════════════
    # SHEET 3  —  Detailed_Metrics
    # ═══════════════════════════════════════════════════════════════════════
    ws_det = wb.create_sheet("Detailed_Metrics")
    metric_order = ["Source_Experiment","Domain_Type","Analysis","Model",
                    "Accuracy","F1","AUC","AUC_CI_lo","AUC_CI_hi","MCC"]
    det_cols = [c for c in metric_order if c in raw_df.columns]
    detailed  = raw_df[det_cols].sort_values(
        ["Source_Experiment","Domain_Type","Analysis","AUC"],
        ascending=[True,True,True,False]
    )
    for r in dataframe_to_rows(detailed, index=False, header=True):
        ws_det.append(r)

    format_sheet_base(ws_det)

    # FIX 3 — correct MCC colouring
    col_map_det = {name: i+1 for i, name in enumerate(det_cols)}
    metric_cols_det = [col_map_det[c] for c in
                       ["Accuracy","F1","AUC","MCC"] if c in col_map_det]
    mcc_cols_det    = [col_map_det["MCC"]] if "MCC" in col_map_det else []
    highlight_best_worst(ws_det, metric_cols_det, mcc_cols_det)

    # Domain colour banding
    if "Analysis" in det_cols:
        an_col = det_cols.index("Analysis") + 1
        colour_domain_rows(ws_det, an_col)

    # ═══════════════════════════════════════════════════════════════════════
    # SHEET 4  —  AUC_Confidence_Intervals  (FIX 5)
    # ═══════════════════════════════════════════════════════════════════════
    ws_ci = wb.create_sheet("AUC_Confidence_Intervals")
    ci_df = build_ci_summary(raw_df)
    for r in dataframe_to_rows(ci_df, index=False, header=True):
        ws_ci.append(r)

    format_sheet_base(ws_ci)

    # Highlight CI_Width column: wide = red, narrow = green
    if "CI_Width" in ci_df.columns:
        ci_col_map  = {name: i+1 for i, name in enumerate(ci_df.columns)}
        w_col       = ci_col_map.get("CI_Width")
        auc_col     = ci_col_map.get("AUC")
        if w_col:
            for r in range(2, ws_ci.max_row + 1):
                w = ws_ci.cell(row=r, column=w_col).value
                if isinstance(w, float):
                    if w > 0.20:
                        ws_ci.cell(row=r, column=w_col).fill = WORST_FILL
                        ws_ci.cell(row=r, column=w_col).font = WORST_FONT
                    elif w < 0.08:
                        ws_ci.cell(row=r, column=w_col).fill = BEST_FILL
                        ws_ci.cell(row=r, column=w_col).font = BEST_FONT

        # Domain banding
        if "Analysis" in ci_df.columns:
            an_col = list(ci_df.columns).index("Analysis") + 1
            colour_domain_rows(ws_ci, an_col)

    # ═══════════════════════════════════════════════════════════════════════
    # SHEET 5  —  Generalisation_Verdict
    # ═══════════════════════════════════════════════════════════════════════
    ws_gen = wb.create_sheet("Generalisation_Verdict")
    gen_df = build_generalisation_verdict(raw_df)
    if not gen_df.empty:
        for r in dataframe_to_rows(gen_df, index=False, header=True):
            ws_gen.append(r)
        format_sheet_base(ws_gen)

        gap_col_idx = list(gen_df.columns).index("Generalisation_Gap") + 1
        for r in range(2, ws_gen.max_row + 1):
            gap = ws_gen.cell(row=r, column=gap_col_idx).value
            if isinstance(gap, float):
                if gap > 0.25:
                    for c in range(1, ws_gen.max_column + 1):
                        ws_gen.cell(row=r, column=c).fill = WORST_FILL
                elif gap < 0.12:
                    for c in range(1, ws_gen.max_column + 1):
                        ws_gen.cell(row=r, column=c).fill = BEST_FILL
                else:
                    for c in range(1, ws_gen.max_column + 1):
                        ws_gen.cell(row=r, column=c).fill = WARN_FILL

    # ═══════════════════════════════════════════════════════════════════════
    # SHEET 6  —  Statistical_Analysis  (per experiment, stratified)
    # ═══════════════════════════════════════════════════════════════════════
    ws_stat = wb.create_sheet("Statistical_Analysis")
    numeric_cols = [c for c in ["Accuracy","F1","AUC","MCC"] if c in raw_df.columns]
    stats = (
        raw_df
        .groupby(["Source_Experiment","Domain_Type"])[numeric_cols]
        .agg(["mean","std","min","max"])
        .reset_index()
    )
    stats.columns = [
        "_".join(col).strip().rstrip("_") if col[1] else col[0]
        for col in stats.columns.values
    ]
    for r in dataframe_to_rows(stats, index=False, header=True):
        ws_stat.append(r)
    format_sheet_base(ws_stat)

    # ═══════════════════════════════════════════════════════════════════════
    # SHEET 7  —  Comparison_Matrix
    # ═══════════════════════════════════════════════════════════════════════
    ws_comp = wb.create_sheet("Comparison_Matrix")
    ws_comp.append(["NOTE: Values are mean AUC averaged across all experiments "
                    "for this Analysis×Model pair. Stratify by Domain_Type "
                    "for scientific validity."])
    ws_comp["A1"].font = Font(italic=True, color="595959")
    ws_comp.append([])

    try:
        comp = (
            raw_df
            .pivot_table(index="Model", columns="Analysis",
                         values="AUC", aggfunc="mean")
            .reset_index()
        )
        comp.fillna("", inplace=True)
        for r in dataframe_to_rows(comp, index=False, header=True):
            ws_comp.append(r)
    except Exception as e:
        ws_comp.append([f"Error building matrix: {e}"])

    format_sheet_base(ws_comp)

    # ═══════════════════════════════════════════════════════════════════════
    # SHEET 8  —  Rankings
    # ═══════════════════════════════════════════════════════════════════════
    ws_rank = wb.create_sheet("Rankings")
    ranked  = raw_df.sort_values("AUC", ascending=False).reset_index(drop=True)
    ranked["Global_Rank"] = ranked.index + 1
    rank_cols = ["Global_Rank","Source_Experiment","Domain_Type",
                 "Analysis","Model","AUC","AUC_CI_lo","AUC_CI_hi","Accuracy","F1","MCC"]
    rank_cols = [c for c in rank_cols if c in ranked.columns]
    for r in dataframe_to_rows(ranked[rank_cols], index=False, header=True):
        ws_rank.append(r)
    format_sheet_base(ws_rank)

    # FIX 3 — MCC colouring in Rankings
    rank_col_map = {name: i+1 for i, name in enumerate(rank_cols)}
    metric_cols_rank = [rank_col_map[c] for c in
                        ["AUC","Accuracy","F1","MCC"] if c in rank_col_map]
    mcc_cols_rank    = [rank_col_map["MCC"]] if "MCC" in rank_col_map else []
    highlight_best_worst(ws_rank, metric_cols_rank, mcc_cols_rank)

    if "Analysis" in rank_cols:
        an_col = rank_cols.index("Analysis") + 1
        colour_domain_rows(ws_rank, an_col)

    # ═══════════════════════════════════════════════════════════════════════
    # SHEET 9  —  Raw_Data
    # ═══════════════════════════════════════════════════════════════════════
    ws_raw = wb.create_sheet("Raw_Data")
    for r in dataframe_to_rows(raw_df, index=False, header=True):
        ws_raw.append(r)
    format_sheet_base(ws_raw)

    # Thick top border on each experiment boundary
    exp_col_idx = (list(raw_df.columns).index("Source_Experiment") + 1
                   if "Source_Experiment" in raw_df.columns else 1)
    last_exp = None
    for r in range(2, ws_raw.max_row + 1):
        curr = ws_raw.cell(row=r, column=exp_col_idx).value
        if curr and curr != last_exp and last_exp is not None:
            for c in range(1, ws_raw.max_column + 1):
                cell = ws_raw.cell(row=r, column=c)
                cell.border = Border(
                    top    = THICK_TOP,
                    bottom = THIN,
                    left   = THIN,
                    right  = THIN,
                )
        last_exp = curr

    # ═══════════════════════════════════════════════════════════════════════
    # SHEET 10  —  Visual_Summary  (FIX 6)
    # ═══════════════════════════════════════════════════════════════════════
    ws_vis = wb.create_sheet("Visual_Summary")
    if not vis_df.empty:
        for r in dataframe_to_rows(vis_df, index=False, header=True):
            ws_vis.append(r)
        # Flag UNKNOWN type rows
        type_col = list(vis_df.columns).index("Type") + 1
        for r in range(2, ws_vis.max_row + 1):
            v = ws_vis.cell(row=r, column=type_col).value or ""
            if "UNKNOWN" in v:
                for c in range(1, ws_vis.max_column + 1):
                    ws_vis.cell(row=r, column=c).fill = DUPE_FILL
                    ws_vis.cell(row=r, column=c).font = DUPE_FONT
    else:
        ws_vis.append(["Source_Experiment","Image_Name","Type","Auto_Generated_Summary"])
        ws_vis.append(["None","None","None","No images found."])
    format_sheet_base(ws_vis)

    # ═══════════════════════════════════════════════════════════════════════
    # SAVE
    # ═══════════════════════════════════════════════════════════════════════
    log.info(f"\nSaving → {OUTPUT_FILE}")
    wb.save(OUTPUT_FILE)

    # ─── Final summary ───────────────────────────────────────────────────────
    log.info("\n" + "=" * 68)
    log.info("REPORT COMPLETE")
    log.info(f"  Experiments loaded  : {raw_df['Source_Experiment'].nunique()}")
    log.info(f"  Rows loaded         : {len(raw_df)}")
    log.info(f"  Missing CSV files   : {(~load_log_df['CSV_Found']).sum()}")
    log.info(f"  Duplicate pairs     : {len(dupe_df)}")
    log.info(f"  Output              : {OUTPUT_FILE}")
    log.info("=" * 68)
    log.info("\nACTION ITEMS:")
    log.info("  1. Open 'Data_Integrity' sheet FIRST — check duplicates.")
    log.info("  2. Read 'Overview_Summary' by Domain_Type, NOT the global mean.")
    log.info("  3. Use 'AUC_Confidence_Intervals' to spot small-sample rows (⚠ WIDE).")
    log.info("  4. Use 'Generalisation_Verdict' to see which models truly transfer.")
    log.info("  5. All negative MCC cells are now correctly flagged RED — not green.")


if __name__ == "__main__":
    build_report()
